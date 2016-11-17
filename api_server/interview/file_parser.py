#! -*- coding: utf-8 -*-

from openpyxl import load_workbook


def file_parser(ext_name, content):
    '''
    Parse candidate data from csv and xlsx files.
    Caution: candidate id is not generated here.
    '''
    candidate_list = []
    if ext_name == 'csv':
        text = content.read().decode('utf-8')
        spamreader = '\n'.join(text.split('\r\n')).split('\n')
        line_counter = 0
        for raw_row in spamreader:
            line_counter += 1
            if line_counter <= 1:
                row = raw_row.split(',')
                if len(row) < 3 or row[0] != u'姓名' or \
                        row[1] != u'邮箱' or row[2] != u'手机号':
                    return None
                continue
            row = raw_row.split(',')
            if row[0] == '' or row[0] is None:
                continue
            candidate_data = {
                    'id': '',
                    'name': row[0],
                    'email': row[1],
                    'phone': row[2],
                    'status': u'未面试',
                    'roomId': '',
                    'record': {
                        'video': 0,
                        'board': 0,
                        'chat': 0,
                        'code': 0,
                        'report': 9
                    }
                }
            candidate_list.append(candidate_data)
            # print (candidate_list)
    elif ext_name == 'xlsx':
        wb = load_workbook(content)
        # wb = load_workbook(filename=r'example1.xlsx')
        sheets = wb.get_sheet_names()
        sheet0 = sheets[0]
        ws = wb.get_sheet_by_name(sheet0)
        rows = ws.rows
        line_counter = 0
        for row in rows:
            line_counter += 1
            if line_counter <= 2:
                if line_counter == 2:
                    line = [col.value for col in row]
                    if len(line) < 3 or line[0] != u'姓名' or \
                       line[1] != u'邮箱' or line[2] != u'手机号':
                        return None
                continue
            line = [col.value for col in row]
            if line[0] == '' or line[0] is None:
                continue
            candidate_data = {
                'id': '',
                'name': line[0],
                'email': line[1],
                'phone': line[2],
                'status': u'未面试',
                'roomId': '',
                'record': {
                    'video': 0,
                    'board': 0,
                    'chat': 0,
                    'code': 0,
                    'report': 0
                }
            }
            candidate_list.append(candidate_data)
    else:
        print ("Unknown file format.")
        return None
    return candidate_list

if __name__ == "__main__":
    file_parser('example1.xlsx')
    file_parser('example2.csv')
